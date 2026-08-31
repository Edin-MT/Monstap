"""
exporter.py
-----------
Eksport obrađenih podataka u novi Excel fajl sa jednim ili više sheet-ova
i osnovnim formatiranjem: podebljano/obojeno zaglavlje, zamrznut red
zaglavlja, auto-filter, brojni format i automatska širina kolona.
"""

from __future__ import annotations

from pathlib import Path
from typing import Dict

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center")

MIN_COLUMN_WIDTH = 8
MAX_COLUMN_WIDTH = 60


def export_workbook(output_path: str, sheets: Dict[str, pd.DataFrame], output_cfg: dict) -> None:
    """Upisuje jedan ili više DataFrame-ova u novi .xlsx fajl, po jedan sheet svaki.

    Args:
        output_path: putanja izlaznog .xlsx fajla (roditeljski folder se kreira ako ne postoji).
        sheets: rečnik {ime_sheeta: DataFrame}; redosled unosa se čuva u izlaznom fajlu.
        output_cfg: 'output' sekcija konfiguracije - koristi se za formatiranje
            ('number_format', 'freeze_header', 'autofilter').

    Raises:
        ValueError: ako 'sheets' nema nijedan unos.
    """
    if not sheets:
        raise ValueError("Nema podataka za eksport (prazna lista sheet-ova).")

    out_path = Path(output_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)

    number_format = output_cfg.get("number_format", "#,##0.00")
    freeze_header = output_cfg.get("freeze_header", True)
    autofilter = output_cfg.get("autofilter", True)

    # Excel ograničava imena sheet-ova na 31 karakter - unapred računamo
    # "bezbedna" imena da bismo ih koristili i za pisanje i za formatiranje.
    safe_names = {name: name[:31] for name in sheets}

    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        for name, df in sheets.items():
            df.to_excel(writer, sheet_name=safe_names[name], index=False)

        for name, df in sheets.items():
            worksheet = writer.sheets[safe_names[name]]
            _format_sheet(
                worksheet,
                df,
                number_format=number_format,
                freeze_header=freeze_header,
                autofilter=autofilter,
            )


def _format_sheet(
    worksheet: Worksheet,
    df: pd.DataFrame,
    number_format: str,
    freeze_header: bool,
    autofilter: bool,
) -> None:
    """Primenjuje osnovno formatiranje na jedan sheet (zaglavlje, format, širine)."""
    n_rows, n_cols = df.shape
    if n_cols == 0:
        return

    # Zaglavlje: podebljano, obojeno, centrirano.
    for col_idx in range(1, n_cols + 1):
        cell = worksheet.cell(row=1, column=col_idx)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT

    # Brojni format za numeričke (ne-bool) kolone.
    for col_idx, column_name in enumerate(df.columns, start=1):
        series = df[column_name]
        if pd.api.types.is_numeric_dtype(series) and not pd.api.types.is_bool_dtype(series):
            for row_idx in range(2, n_rows + 2):
                worksheet.cell(row=row_idx, column=col_idx).number_format = number_format

    # Automatska širina kolone na osnovu najdužeg sadržaja (zaglavlje ili vrednost).
    for col_idx, column_name in enumerate(df.columns, start=1):
        max_len = len(str(column_name))
        if n_rows > 0:
            max_len = max(max_len, df[column_name].astype(str).map(len).max())
        width = min(max(max_len + 2, MIN_COLUMN_WIDTH), MAX_COLUMN_WIDTH)
        worksheet.column_dimensions[get_column_letter(col_idx)].width = width

    if freeze_header:
        worksheet.freeze_panes = "A2"

    if autofilter and n_rows > 0:
        last_col_letter = get_column_letter(n_cols)
        worksheet.auto_filter.ref = f"A1:{last_col_letter}{n_rows + 1}"
